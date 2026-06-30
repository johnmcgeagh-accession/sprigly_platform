#!/usr/bin/env python3
"""
generate_calendar.py  —  CSV (+optional config) → branded 3-tab Excel content calendar
Usage: python generate_calendar.py <csv_path> [--config <path>] [--out <path>]
Requires: pip install openpyxl
"""

import argparse, csv, json, calendar, math, re, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour constants ─────────────────────────────────────────────────────────
SLATE    = "334155"
INK      = "1E293B"
WHITE    = "FFFFFF"
GREY_BG  = "F7F8F9"
GREY_TXT = "94A3B8"
BORDER_C = "D5D9DE"
AMBER    = "F59E0B"
AMBER_BG = "FFF6E5"
CORAL    = "FF6F62"

# Patched for worker deployment: palette.json lives alongside this script.
# (Original skill had parent.parent to reach the skill root; here it's just parent.)
SKILL_DIR = Path(__file__).resolve().parent

MONTH_NAMES = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
WEEKDAYS = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

# ── Style helpers ────────────────────────────────────────────────────────────

def hf(hex_str):
    return PatternFill("solid", fgColor=hex_str.lstrip("#"))

def bdr(colour=BORDER_C):
    s = Side(style="thin", color=colour.lstrip("#"))
    return Border(left=s, right=s, top=s, bottom=s)

def fnt(size=9, bold=False, italic=False, colour=INK):
    return Font(name="Arial", size=size, bold=bold, italic=italic,
                color=colour.lstrip("#"))

def aln(h="left", v="top", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(ws, row, col, value="", f=None, fill=None, al=None, bd=None):
    """Set cell value and optional styles."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if f:    c.font      = f
    if fill: c.fill      = fill
    if al:   c.alignment = al
    if bd:   c.border    = bd
    return c

def mc(ws, r1, c1, r2, c2, value="", f=None, fill=None, al=None):
    """Merge cells and style the top-left."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return sc(ws, r1, c1, value=value, f=f, fill=fill, al=al)

# ── Row height estimator ─────────────────────────────────────────────────────

def estimate_lines(text, col_width):
    """Count visual lines for text in a column of given character width."""
    if not text or col_width <= 0:
        return 1
    effective = col_width * 1.05
    lines = 0
    for seg in str(text).split("\n"):
        lines += math.ceil((len(seg) + 1) / effective)
    return max(lines, 1)

def row_height(cells_and_widths, min_h=40, max_h=560):
    """Compute row height from list of (text, col_width) pairs."""
    h = min_h
    for text, width in cells_and_widths:
        lines = estimate_lines(text, width)
        h = max(h, lines * 11.5 + 6)
    return min(h, max_h)

# ── Date parser ──────────────────────────────────────────────────────────────

def parse_day(date_str, month):
    """Extract the day-of-month integer from a Date cell value."""
    if not date_str:
        return None
    s = str(date_str).strip()
    # YYYY-MM-DD
    m = re.match(r"\d{4}-(\d{2})-(\d{2})", s)
    if m and int(m.group(1)) == month:
        return int(m.group(2))
    # "D Mon" / "DD Mon"
    m = re.match(r"^(\d{1,2})\s+\w+", s)
    if m:
        return int(m.group(1))
    # bare number
    m = re.match(r"^(\d{1,2})$", s)
    if m:
        return int(m.group(1))
    return None

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Generate a branded Excel content calendar from a CSV plan."
    )
    ap.add_argument("csv_path",        help="Path to the content-plan CSV (must start YYYY-MM_)")
    ap.add_argument("--config",        help="Path to calendar-config.json", default=None)
    ap.add_argument("--out",           help="Output directory or full .xlsx path", default=None)
    args = ap.parse_args()

    csv_path = Path(args.csv_path)

    # Derive year / month from filename prefix YYYY-MM_
    mm = re.match(r"^(\d{4})-(\d{2})_", csv_path.stem)
    if not mm:
        sys.exit(f"Error: filename must start with YYYY-MM_  (got: {csv_path.stem})")
    year, month = int(mm.group(1)), int(mm.group(2))
    month_name = MONTH_NAMES[month - 1]

    # Load config
    config = {}
    if args.config:
        with open(args.config, encoding="utf-8") as f:
            config = json.load(f)

    raw_slug = csv_path.stem[8:].replace("-", " ").replace("_", " ").strip()
    client   = config.get("client", raw_slug.title() or "Client")
    contact  = config.get("contact", "the client")
    cat_cfg  = config.get("categories", {})

    # Load palette
    try:
        palette = json.loads((SKILL_DIR / "palette.json").read_text())
    except Exception:
        palette = ["FFD7D1","FCE7C3","D5E5F2","DBEEDD","E6DEF2",
                   "CFE9E4","DDE6F0","F6DDEA","FBE2CE","E6E9ED"]

    # Load CSV
    posts = []
    csv_headers = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        csv_headers = list(reader.fieldnames or [])
        for row in reader:
            if any(v.strip() for v in row.values() if v):
                posts.append(row)

    # Build category → colour map (config values first, then auto from palette)
    cat_colours = dict(cat_cfg)
    cat_order   = list(cat_cfg.keys())
    pal_idx     = len(cat_colours)

    for post in posts:
        cat = (post.get("Category") or "").strip()
        if cat and cat not in cat_colours:
            cat_colours[cat] = palette[pal_idx % len(palette)]
            cat_order.append(cat)
            pal_idx += 1

    # Day → post lookup
    day_posts = {}
    for post in posts:
        day = parse_day(post.get("Date", ""), month)
        if day is not None:
            day_posts[day] = post

    # Category counts for legend
    cat_counts = {}
    for post in posts:
        cat = (post.get("Category") or "").strip()
        if cat:
            cat_counts[cat] = cat_counts.get(cat, 0) + 1

    # Locate editable columns by suffix
    def by_suffix(headers, suffix):
        return next((h for h in headers if h and h.endswith(suffix)), None)

    amended_hdr = by_suffix(csv_headers, "Amended Caption") or f"{contact}'s Amended Caption"
    notes_hdr   = by_suffix(csv_headers, "Notes / Questions") or f"{contact}'s Notes / Questions"
    sprigly_notes_hdr = next((h for h in csv_headers if "Sprigly Notes" in h),
                              "Sprigly Notes (context for client)")

    # Post details column definitions: (csv_key, display_label, width, align, editable)
    #   align: 'c' = centre, 'w' = wrap-left, 'e' = editable
    DETAIL_COLS = [
        ("Date",                                              "Date",                  8,  "c", False),
        ("Day",                                               "Day",                   6,  "c", False),
        ("Post Title / Theme",                                "Post Title / Theme",    26, "w", False),
        ("Category",                                          "Category",              17, "w", False),
        ("Pillar",                                            "Pillar",                22, "w", False),
        ("Format",                                            "Format",                18, "w", False),
        ("Posting Time",                                      "Posting Time",          11, "c", False),
        ("Who Posts",                                         "Who Posts",             16, "w", False),
        ("Competitor Insight (why this was recommended)",
         "Competitor Insight (why this was recommended)",     52, "w", False),
        ("Sprigly Draft Caption",                             "Sprigly Draft Caption", 56, "w", False),
        (sprigly_notes_hdr,                                   sprigly_notes_hdr,       40, "w", False),
        (amended_hdr,                                         amended_hdr,             50, "e", True),
        (notes_hdr,                                           notes_hdr,               32, "e", True),
    ]

    # ════════════════════════════════════════════════════════════════════════
    # Build workbook
    # ════════════════════════════════════════════════════════════════════════
    wb = Workbook()
    wb.remove(wb.active)

    # ────────────────────────────────────────────────────────────────────────
    # TAB 1 — Calendar
    # ────────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Calendar")
    ws1.sheet_view.showGridLines = False

    for ci in range(1, 8):
        ws1.column_dimensions[get_column_letter(ci)].width = 24

    # Row 1: title
    ws1.row_dimensions[1].height = 28
    mc(ws1, 1, 1, 1, 7,
       value=f"{client} — Content calendar",
       f=fnt(size=20, bold=True, colour=SLATE),
       al=aln("left", "center"))

    # Row 2: subtitle
    ws1.row_dimensions[2].height = 18
    mc(ws1, 2, 1, 2, 7,
       value=(f"{month_name} {year}  ·  colours show the post type  ·  "
              f'full captions on the "Post details" tab'),
       f=fnt(size=10, italic=True, colour=GREY_TXT),
       al=aln("left", "center"))

    # Row 3: blank spacer
    ws1.row_dimensions[3].height = 8

    # Row 4: weekday headers
    ws1.row_dimensions[4].height = 22
    for ci, name in enumerate(WEEKDAYS, start=1):
        sc(ws1, 4, ci, name,
           f=fnt(size=8, bold=True, colour=WHITE),
           fill=hf(SLATE),
           al=aln("center", "center"))

    ws1.freeze_panes = "A5"

    # Month grid
    cal_grid = calendar.Calendar(firstweekday=0)
    weeks = cal_grid.monthdayscalendar(year, month)
    cell_bdr = bdr()

    for wi, week in enumerate(weeks):
        r_num  = 5 + wi * 2
        r_cont = 5 + wi * 2 + 1
        ws1.row_dimensions[r_num].height  = 16
        ws1.row_dimensions[r_cont].height = 96

        for ci, day in enumerate(week, start=1):
            if day == 0:
                sc(ws1, r_num,  ci, fill=hf(GREY_BG), bd=cell_bdr)
                sc(ws1, r_cont, ci, fill=hf(GREY_BG), bd=cell_bdr)
            else:
                post = day_posts.get(day)
                cat  = (post.get("Category") or "").strip() if post else None
                fc   = cat_colours.get(cat, WHITE) if cat else WHITE

                sc(ws1, r_num, ci, value=day,
                   f=fnt(size=8, colour=SLATE),
                   fill=hf(fc),
                   al=aln("left", "top"),
                   bd=cell_bdr)

                if post:
                    title = post.get("Post Title / Theme") or ""
                    catv  = post.get("Category") or ""
                    timev = post.get("Posting Time") or ""
                    fmtv  = post.get("Format") or ""
                    whov  = post.get("Who Posts") or ""
                    body  = f"{title}\n{catv} · {timev}\n{fmtv}\n→ {whov}"
                    sc(ws1, r_cont, ci, value=body,
                       f=fnt(size=8.5, colour=SLATE),
                       fill=hf(fc),
                       al=aln("left", "top", wrap=True),
                       bd=cell_bdr)
                else:
                    sc(ws1, r_cont, ci, fill=hf(WHITE), bd=cell_bdr)

    # Legend
    leg_row = 5 + len(weeks) * 2 + 2
    ws1.row_dimensions[leg_row].height = 16
    sc(ws1, leg_row, 1, "Post types",
       f=fnt(size=8, bold=True, colour=SLATE),
       al=aln("left", "center"))

    for li, cat in enumerate(cat_order):
        if cat not in cat_colours:
            continue
        lr = leg_row + 1 + li
        ws1.row_dimensions[lr].height = 16
        count = cat_counts.get(cat, 0)
        sc(ws1, lr, 1, fill=hf(cat_colours[cat]))
        ws1.merge_cells(start_row=lr, start_column=2, end_row=lr, end_column=4)
        sc(ws1, lr, 2, value=f"{cat} ({count})",
           f=fnt(size=8.5, colour=SLATE),
           al=aln("left", "center"))

    # ────────────────────────────────────────────────────────────────────────
    # TAB 2 — Post details
    # ────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Post details")

    # Column widths
    for ci, (_, _, width, _, _) in enumerate(DETAIL_COLS, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = width

    # Header row
    ws2.row_dimensions[1].height = 40
    for ci, (_, label, _, align_type, editable) in enumerate(DETAIL_COLS, start=1):
        if editable:
            display = f"✏️ {label}"
            h_fill  = hf(AMBER)
            h_font  = fnt(size=9, bold=True, colour=INK)
        else:
            display = label
            h_fill  = hf(SLATE)
            h_font  = fnt(size=9, bold=True, colour=WHITE)
        sc(ws2, 1, ci, display,
           f=h_font, fill=h_fill,
           al=aln("center", "center", wrap=True))

    ws2.freeze_panes = "D2"

    # Data rows
    for ri, post in enumerate(posts, start=2):
        # Estimate row height from non-editable wrap columns
        pairs = []
        for csv_key, _, width, align_type, editable in DETAIL_COLS:
            if not editable and align_type == "w":
                val = post.get(csv_key) or ""
                pairs.append((val, width))
        ws2.row_dimensions[ri].height = row_height(pairs)

        cat  = (post.get("Category") or "").strip()
        cat_fc = cat_colours.get(cat, WHITE) if cat else WHITE

        for ci, (csv_key, _, width, align_type, editable) in enumerate(DETAIL_COLS, start=1):
            if editable:
                sc(ws2, ri, ci, value="",
                   f=fnt(size=9, colour=INK),
                   fill=hf(AMBER_BG),
                   al=aln("left", "top", wrap=True))
            else:
                val      = post.get(csv_key) or ""
                cell_fill = hf(cat_fc) if csv_key == "Category" else hf(WHITE)
                h_align   = "center" if align_type == "c" else "left"
                sc(ws2, ri, ci, value=val,
                   f=fnt(size=9, colour=SLATE),
                   fill=cell_fill,
                   al=aln(h_align, "top", wrap=True))

    # ────────────────────────────────────────────────────────────────────────
    # TAB 3 — How to use
    # ────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("How to use")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 96

    def how(row, text, f=None, al=None, height=None):
        if height:
            ws3.row_dimensions[row].height = height
        if text is not None:
            sc(ws3, row, 2, text, f=f, al=al)

    contact_title = contact.title() if contact != "the client" else "Your contact"
    r = 1

    how(r, f"{client} — Content calendar, {month_name} {year}",
        f=fnt(size=12, bold=True, colour=SLATE),
        al=aln("left", "center"), height=26)
    r += 1
    how(r, None, height=6); r += 1

    how(r, "How this works",
        f=fnt(size=10, bold=True, colour=SLATE),
        al=aln("left", "top"), height=20); r += 1

    how(r, ("The Calendar tab shows the month at a glance. "
            "Colours indicate post type — the key is at the bottom. "
            "Use it to see the shape of the month before you look at the detail."),
        f=fnt(size=9, colour=INK),
        al=aln("left", "top", wrap=True), height=36); r += 1

    how(r, ("The Post details tab has the full plan: one row per post, the draft caption, "
            "why we suggested that approach (Competitor Insight column), and our notes. "
            "That's the tab you work in."),
        f=fnt(size=9, colour=INK),
        al=aln("left", "top", wrap=True), height=36); r += 1

    how(r, None, height=8); r += 1

    how(r, "Your bit",
        f=fnt(size=10, bold=True, colour=CORAL),
        al=aln("left", "top"), height=20); r += 1

    how(r, (f"The two amber columns are yours. \"{contact_title}'s Amended Caption\" "
            f"is where you rewrite a draft if you’d say it differently. "
            f"\"{contact_title}'s Notes / Questions\" is for anything "
            "you want to flag, query, or change about the post itself."),
        f=fnt(size=9, colour=INK),
        al=aln("left", "top", wrap=True), height=48); r += 1

    how(r, "Leave a row blank if the draft's good. Blank means happy with it.",
        f=fnt(size=9, bold=True, colour=INK),
        al=aln("left", "top"), height=20); r += 1

    how(r, None, height=8); r += 1

    how(r, "When you send it back",
        f=fnt(size=10, bold=True, colour=CORAL),
        al=aln("left", "top"), height=20); r += 1

    how(r, ("We compare your edits against our drafts and feed the differences into "
            "how your agent writes going forward. The more you rewrite in your own words, "
            "the closer the next plan will sound like you. "
            "One round of edits is usually enough to make a noticeable difference."),
        f=fnt(size=9, colour=INK),
        al=aln("left", "top", wrap=True), height=52); r += 1

    how(r, None, height=8); r += 1

    how(r, ("Some rows may carry placeholder dates. "
            "If a post moves, shift all the posts in that week together to keep the sequence."),
        f=fnt(size=9, italic=True, colour=GREY_TXT),
        al=aln("left", "top", wrap=True), height=28); r += 1

    how(r, "No formulas anywhere. Just save and send back the file.",
        f=fnt(size=9, italic=True, colour=GREY_TXT),
        al=aln("left", "top"), height=16); r += 1

    how(r, None, height=8); r += 1

    how(r, "Sprigly · sprigly.co.uk",
        f=fnt(size=9, colour=GREY_TXT),
        al=aln("left", "bottom"), height=16)

    # ── Save ─────────────────────────────────────────────────────────────────
    out_name = f"{client} — Content calendar - {month_name} {year}.xlsx"

    if args.out:
        out_p = Path(args.out)
        out_path = (out_p / out_name) if out_p.is_dir() else out_p
    else:
        out_path = csv_path.parent / out_name

    wb.save(str(out_path))
    print(f"Saved:      {out_path}")
    print(f"Posts:      {len(posts)}")
    print(f"Categories: {len(cat_colours)} ({', '.join(cat_order)})")
    print(f"Weeks:      {len(weeks)}")


if __name__ == "__main__":
    main()
