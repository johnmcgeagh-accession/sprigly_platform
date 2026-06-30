#!/usr/bin/env python3
"""
extract_edits.py  —  edited xlsx → diff JSON for voice feedback
Usage: python extract_edits.py <edited_xlsx> [--config <path>] [--out <path>]
Requires: pip install openpyxl
"""

import argparse, json, re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

TAB_NAME = "Post details"


def col_index_by_suffix(headers, suffix):
    """Return 0-based column index of the first header ending with suffix."""
    for i, h in enumerate(headers):
        if h and str(h).endswith(suffix):
            return i
    return None


def col_index_exact(headers, name):
    """Return 0-based column index of the first exact header match."""
    for i, h in enumerate(headers):
        if h and str(h) == name:
            return i
    return None


def col_index_contains(headers, fragment):
    """Return 0-based column index of the first header containing fragment."""
    for i, h in enumerate(headers):
        if h and fragment in str(h):
            return i
    return None


def parse_month_from_filename(stem):
    """
    Try to extract YYYY-MM from various filename patterns:
      - "Ivy — Content calendar - July 2026"  →  "2026-07"
      - "2026-07_ivy-t-..."                   →  "2026-07"
    """
    # Pattern: "Month YYYY" at the end
    m = re.search(r"-\s*([A-Za-z]+)\s+(\d{4})", stem)
    if m:
        try:
            dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%B %Y")
            return dt.strftime("%Y-%m")
        except ValueError:
            pass
    # Pattern: YYYY-MM_ prefix
    m = re.match(r"^(\d{4}-\d{2})_", stem.replace(" ", "_"))
    if m:
        return m.group(1)
    return "unknown"


def cell_val(ws, row_idx, col_idx):
    """Return stripped string value of a cell (0-based col index → 1-based openpyxl)."""
    if col_idx is None:
        return ""
    v = ws.cell(row=row_idx, column=col_idx + 1).value
    return str(v).strip() if v is not None else ""


def main():
    ap = argparse.ArgumentParser(
        description="Extract client caption edits from a returned content-calendar xlsx."
    )
    ap.add_argument("edited_xlsx", help="Path to the returned (edited) .xlsx workbook")
    ap.add_argument("--config",    help="Path to calendar-config.json", default=None)
    ap.add_argument("--out",       help="Output .json path", default=None)
    args = ap.parse_args()

    xlsx_path = Path(args.edited_xlsx)

    # Load config
    config = {}
    if args.config:
        with open(args.config, encoding="utf-8") as f:
            config = json.load(f)

    client  = config.get("client", "")
    contact = config.get("contact", "the client")

    # Derive month from filename
    month_key = parse_month_from_filename(xlsx_path.stem)

    # Load workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if TAB_NAME not in wb.sheetnames:
        raise ValueError(
            f"No '{TAB_NAME}' tab found in {xlsx_path.name}. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[TAB_NAME]

    # Read headers from row 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Locate columns
    amended_col  = col_index_by_suffix(headers, "Amended Caption")
    notes_col    = col_index_by_suffix(headers, "Notes / Questions")
    draft_col    = col_index_by_suffix(headers, "Sprigly Draft Caption")
    date_col     = col_index_exact(headers, "Date")
    title_col    = col_index_exact(headers, "Post Title / Theme")
    cat_col      = col_index_exact(headers, "Category")
    pillar_col   = col_index_exact(headers, "Pillar")

    if amended_col is None:
        raise ValueError("Could not locate an 'Amended Caption' column in Post details tab.")
    if notes_col is None:
        raise ValueError("Could not locate a 'Notes / Questions' column in Post details tab.")

    # Walk data rows (skip header row 1)
    edits = []
    total_posts = 0

    for row_idx in range(2, ws.max_row + 1):
        # Skip fully empty rows
        if not any(ws.cell(row=row_idx, column=c).value
                   for c in range(1, ws.max_column + 1)):
            continue

        total_posts += 1

        amended = cell_val(ws, row_idx, amended_col)
        notes   = cell_val(ws, row_idx, notes_col)

        if not amended and not notes:
            continue  # no edit

        edits.append({
            "date":         cell_val(ws, row_idx, date_col),
            "post_title":   cell_val(ws, row_idx, title_col),
            "category":     cell_val(ws, row_idx, cat_col),
            "pillar":       cell_val(ws, row_idx, pillar_col),
            "sprigly_draft": cell_val(ws, row_idx, draft_col),
            "amended":      amended,
            "notes":        notes,
            "changed":      True,
        })

    result = {
        "client":  client,
        "contact": contact,
        "month":   month_key,
        "edits":   edits,
        "summary": {
            "total_posts": total_posts,
            "edited":      len(edits),
            "edit_rate":   round(len(edits) / total_posts, 2) if total_posts else 0,
        },
    }

    # Output path
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = xlsx_path.parent / f"{xlsx_path.stem}-edits.json"

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"Saved:      {out_path}")
    print(f"Posts:      {total_posts}")
    print(f"Edited:     {len(edits)}")
    print(f"Edit rate:  {result['summary']['edit_rate']:.0%}")


if __name__ == "__main__":
    main()
